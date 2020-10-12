#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Dec 14 09:27:25 2019

@author: Federica LARENO FACCINI

Create excel file as DATABASE of BEHAVIOURAL DATA.
One file for each animal.
The first time it runs for one animal it creates the file (first session). After that, everytime it runs, it adds a new sheet for the new behaviour session.
The ENVELOPE (n by bin of the PSTH by session)is saved in a specific sheet, one column for each session.
The data stored for each session is: trial number, delay of the reward, opening time of the valve (the water appears at the end of the OT), number of licks per trial and each individual lick time.
"""
import extrapy.Behaviour as bv
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
import glob
import matplotlib.pyplot as plt

group = 14
mouse = 6402
experiment = 'Fixed Delay'
session = 'EF-P18'
skiplast = False #True only for RANDOM DELAY!

og_path = fr"D:\F.LARENO.FACCINI\Preliminary Results\Behaviour\Group {group}\{mouse}\{experiment}\{session}\*.lick"
files = glob.glob(og_path)
#extract file name
files = [os.path.basename(i) for i in files]

basepath = fr"D:/F.LARENO.FACCINI/Preliminary Results/Behaviour/Group {group}/{mouse}/{experiment}/{session}"
savedir = r"D:\F.LARENO.FACCINI\Preliminary Results\Behaviour\Database\Prova"

# During training mistakes in the recordings can happen.
# These mistakes can lead to shorter recordings and thus to the creation of multiple files for the same session.
# With the following loop I check if there are multiple files in the same session.
# If so, I concatenate them so that the trial number becomes sequetial and I concatenate also the delays
if len(files)>1:
    licks = bv.concatenate_licks(basepath,skiplast) 
    files = [f.replace(".lick","") for f in files]
    trials = [0] # Initialize counter for trial number
    for idx,file in enumerate(files):
        #recreate path
        param = basepath+'/'+file+".param"
        #Load random delays 
        random = bv.extract_random_delay(param,skiplast,fixed_delay=500) 
        # Format the delay array
        delay = np.asarray([d for d, _ in (random)])
        ot = np.asarray(bv.extract_ot(param,skiplast)) 
        if idx ==0:
            delays = delay
            ots = ot
        else:
            delays = np.concatenate((delays,delay))
            ots = np.concatenate((ots,ot))
        # Format the trial array
        trial = np.asarray([d for _,d in (random)])
        trial[:] = [i+trials[-1] for i in trial]
        trials = np.append(trials,trial)
    trials = np.delete(trials,0)
    new_random = [list(a) for a in zip(delays,trials)]
    if 'Fixed Delay' not in experiment and 'NB' not in session:# and 'T11-2' not in session:
        print('with delays')
        _, licks_by_delay = bv.separate_by_delay(new_random, licks)

# In case there is only one file per session
elif len(files)==1:
    file = files[0].replace(".lick","")
    #recreate paths
    path = basepath+'/'+file+".lick"
    param = basepath+'/'+file+".param"
    # Load lick files and random delays
    licks = bv.load_lickfile(path)
    random = bv.extract_random_delay(param,skiplast,fixed_delay=500)
    if 'Fixed Delay' not in experiment and 'NB' not in session:# and 'T11-2' not in session:
        print('With delays')
        _, licks_by_delay = bv.separate_by_delay(random, licks)
    # Format the lists
    delays = np.asarray([d for d, _ in (random)])
    trials = np.asarray([d for _,d in (random)])
    ots = bv.extract_ot(param,skiplast) 


# ========================================================================================
# Divide licks by trial and get number of licks per trial
# ========================================================================================
licks_by_trial = []
num_licks = []

N_TRIALS = int(max(licks[:,0]))

for i in range(N_TRIALS):
    temp__ = [licks[:,1][j] for j in range(len(licks)) if licks[:,0][j]==i+1]
    num_licks.append(len(temp__))

    if len(temp__) > 0:
        licks_by_trial.append(temp__)
    else:
        licks_by_trial.append(np.nan)

licks_by_trial = [np.asarray(x) for x in licks_by_trial]


#Get the length of the longest sequence of licks (the trial with more licks)
for idx,i in enumerate(licks_by_trial):
    if idx==0:
        tem__ = i.size
    else:
        tem__ = np.vstack((tem__,i.size))
max_lick = np.max(tem__)

#convert from list of arrays to 2D array
all_licks = np.zeros((len(licks_by_trial),max_lick))
all_licks[:] = np.nan

for idx,x in enumerate(licks_by_trial):
    l = x.size 
    all_licks[idx,:l] = x


# ========================================================================================
# Extracting the envelope of the PSTH (the n of the PSTH)
# ========================================================================================
if 'P1' in session: # divide NoStim and Stim trials
    nostim_lick, stim_lick = bv.separate_by_condition(licks)
    ns,nsbins,_ = bv.psth_lick(nostim_lick,samp_period=0.01)
    env_df = pd.DataFrame(ns, index=None, columns=[f'{session}_NoStim'])
    bins_df = pd.DataFrame(nsbins, index=None, columns=[f'{session}_NoStim'])
    sn,sbins,_ = bv.psth_lick(stim_lick,samp_period=0.01)
    env_df[f'{session}_Stim'] = sn
    bins_df[f'{session}_Stim'] = sbins
    plt.close('all')
else: # the training has no stim so no need for the division
    n,bins,_ = bv.psth_lick(licks,samp_period=0.01)
    env_df = pd.DataFrame(n, index=None, columns=[f'{session}'])
    bins_df = pd.DataFrame(bins, index=None, columns=[f'{session}'])
    plt.close('all')


if 'Fixed Delay' not in experiment and 'NB' not in session:# and 'T11-2' not in session:
    env_delay, bin_delay = {}, {}
    for index,(key, values) in enumerate(licks_by_delay.items()):
        if len(values) > 0:
            if 'P1' in session:
                nostim_lick_delay, stim_lick_delay = bv.separate_by_condition(values)
                try:
                    nsn_delay,nsbins_delay,_ = bv.psth_lick(nostim_lick_delay,samp_period=0.01)
                except:
                    pass
                try:
                    sn_delay,sbins_delay,_ = bv.psth_lick(stim_lick_delay,samp_period=0.01)
                except:
                    pass
                plt.close('all')
                env_delay[f'{key}_NoStim'] = nsn_delay
                bin_delay[f'{key}_NoStim'] = nsbins_delay
                env_delay[f'{key}_Stim'] = sn_delay
                bin_delay[f'{key}_Stim'] = sbins_delay
            else: 
                n_delay,bins_delay,patches_delay = bv.psth_lick(values,samp_period=0.01)
                plt.close('all')
                env_delay[f'{key}'] = n_delay
                bin_delay[f'{key}'] = bins_delay
    
    env_delay = pd.DataFrame(env_delay)
    bin_delay = pd.DataFrame(bin_delay)

# ========================================================================================
# Create the dataframe
# ========================================================================================
if 'NB' in session:
    delays[:] = np.nan
    ots = np.asarray(ots)
    ots[:] = np.nan
    
cols = ['Trial Number', 'Delay (ms)', 'Opening Time (ms)', 'Number of Licks']
df = pd.DataFrame(zip(trials, delays, ots, num_licks),columns=cols)

# Add the lick times to the df column-wise (so that in the .xlsx each lick time will be stored in its individual cell)
for idx,i in enumerate(all_licks.T):
    df[f'Lick {idx}'] = i[:len(df)] #due to a bug in the logging of trial's parameters in the acquisition software, we have always to delete the last trial of the behaviour. Hence I'm not including the last trial in the df as it's virtually inexistent.

# ========================================================================================
# Create Excel File
# ========================================================================================

#if the file already exists, append a new sheet. To have one file per animal with the different sessions in different sheets
if os.path.isfile(f'{savedir}\Group{group}_{mouse}.xlsx'):
    book = load_workbook(f'{savedir}\Group{group}_{mouse}.xlsx') # THIS CODE IS NOT OPTIMIZED! I'M LOADING THE SAME FILE 3 TIMES I N A ROW (and mixing pandas and openpyxl)!!
    exc = pd.read_excel(f'{savedir}\Group{group}_{mouse}.xlsx', sheet_name='Envelope TOT') # FOR THE MOMENT IT WORKS AND IT'S STILL FAST. BUT ONCE I HAVE TIME I'LL FIX IT (we both know that it will never happen so I'm sorry, it will up to you, dear reader to do it)
    exc_bin = pd.read_excel(f'{savedir}\Group{group}_{mouse}.xlsx', sheet_name='Bins of Envelope')
    if 'P1' in session:
        exc[f'{session}_NoStim'] = ns
        exc[f'{session}_Stim'] = sn
        exc_bin[f'{session}_NoStim'] = nsbins
        exc_bin[f'{session}_Stim'] = sbins
    else:
        exc[f'{session}'] = n
        exc_bin[f'{session}'] = bins

    
    with pd.ExcelWriter(f'{savedir}\Group{group}_{mouse}.xlsx', engine="openpyxl") as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        exc.to_excel(writer, index=False, sheet_name='Envelope TOT')
        exc_bin.to_excel(writer, index=False, sheet_name = 'Bins of Envelope')
        if 'Fixed Delay' not in experiment and 'NB' not in session:# and 'T11-2' not in session:
            for col in env_delay:
                if f'Envelope_{col}' in book.sheetnames:
                    exc_del = pd.read_excel(f'{savedir}\Group{group}_{mouse}.xlsx', sheet_name=f'Envelope_{col}') # FOR THE MOMENT IT WORKS AND IT'S STILL FAST. BUT ONCE I HAVE TIME I'LL FIX IT (we both know that it will never happen so I'm sorry, it will up to you, dear reader to do it)
                    exc_del[f'{session}'] = env_delay[f'{col}']
                    exc_del.to_excel(writer,index=False, sheet_name=f'Envelope_{col}')
                else:
                    exc_del = env_delay[f'{col}']
                    exc_del.rename(f'{session}').to_excel(writer,index=False, sheet_name=f'Envelope_{col}')

            for col in bin_delay:
                if f'Bins_{col}' in book.sheetnames:
                    exc_bin_del = pd.read_excel(f'{savedir}\Group{group}_{mouse}.xlsx', sheet_name=f'Bins_{col}') # FOR THE MOMENT IT WORKS AND IT'S STILL FAST. BUT ONCE I HAVE TIME I'LL FIX IT (we both know that it will never happen so I'm sorry, it will up to you, dear reader to do it)
                    exc_bin_del[f'{session}'] = bin_delay[f'{col}']
                    exc_bin_del.to_excel(writer,index=False, sheet_name=f'Bins_{col}')
                else:
                    bin__ = bin_delay[f'{col}']
                    bin__.rename(f'{session}').to_excel(writer,index=False, sheet_name=f'Bins_{col}')
        df.to_excel(writer, index=False, sheet_name=f"Session {session}")
        
        
        
# creates new excel file if it doesn't exist for that mouse
else:
    with pd.ExcelWriter(f'{savedir}\Group{group}_{mouse}.xlsx', engine="openpyxl") as writer:    
        env_df.to_excel(writer,index=False, sheet_name='Envelope TOT')
        bins_df.to_excel(writer, index=False, sheet_name = 'Bins of Envelope')
        if 'Fixed Delay' not in experiment and 'NB' not in session:# and 'T11-2' not in session:
            for col in env_delay:
                env__ = env_delay[f'{col}']
                env__.rename(f'{session}').to_excel(writer,index=False, sheet_name=f'Envelope_{col}')
            for col in bin_delay:
                bin__ = bin_delay[f'{col}']
                bin__.rename(f'{session}').to_excel(writer,index=False, sheet_name=f'Bins_{col}')
        df.to_excel(writer, index=False, sheet_name=f"Session {session}")

